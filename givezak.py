from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtWidgets import * 
from PyQt5.QtGui import * 
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem
from PyQt5.QtCore import Qt,QTimer
import sys
import os
import openpyxl
import sqlite3

class Give(QMainWindow):
    def __init__(self):
 
        super().__init__()
        

        self.setWindowTitle("Отдать заказ")
        self.setFixedSize(300,150)

        self.add_post = QLineEdit(self)
        self.add_post.move(0, 0)
        self.add_post.resize(300,50)
        self.add_post.setPlaceholderText("Номер заказа")

        self.price = QLineEdit(self)
        self.price.move(0, 50)
        self.price.resize(300,50)
        self.price.setPlaceholderText("Цена")

        self.btn_save = QtWidgets.QPushButton(self)
        self.btn_save.setText("Распечатать текст")
        self.btn_save.move(0,100)
        self.btn_save.setFixedSize(300,50)
        self.btn_save.clicked.connect(self.catch_data)


    def catch_data(self):
        self.db = sqlite3.connect('base.db')
        self.curs = self.db.cursor()
        if len(self.add_post.text()) == 0:
            QMessageBox.warning(self, 'Warning', 'Введите номер заказа')
            return 0
        if len(self.price.text()) == 0:
            QMessageBox.warning(self, 'Warning', 'Введите цену')
            return 0
        self.curs.execute("SELECT * FROM zakaz_stat WHERE number_z = '"+self.add_post.text()+"'")
        self.check = self.curs.fetchall()
        self.curs.execute("SELECT * FROM zakaz WHERE Number = '"+self.add_post.text()+"'")
        self.check1 = self.curs.fetchall()


        

        wb = openpyxl.load_workbook("Price.xlsx")
        wb.copy_worksheet(wb['new'])
        ws = wb['new Copy']
        ws.title = self.check[0][1]        
        ws['B6'] = self.check1[0][1]
        ws['H6'] = self.check1[0][3]
        ws['M6'] = self.price.text()
        ws['M16'] = self.price.text()

        wb.save("Price.xlsx")

        self.curs.execute("DELETE FROM zakaz_stat WHERE number_z = ?",self.add_post.text())
        self.db.commit()
        self.db.close()
        QMessageBox.warning(self, 'Success', 'Данные успешно сохранены')

